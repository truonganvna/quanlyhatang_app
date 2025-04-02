import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import xlsxwriter

# file_path = r'D:\Python\Electric\test.xlsx'
def quan_ly_ha_tang_electric(file_obj):
    df = pd.read_excel(file_obj, skiprows=2)
    df = df.set_index("Địa chỉ")

    # I. Phần dịch vụ
    # Bảng FBS
    df_fbs = df.loc[["AS+ĐL FSB T12", "ĐH-Outdoor T12 (Mới)", "ĐH-Indoor CKTV (Mới)"]]
    df_fbs_frame = pd.DataFrame({
        "Stt": [1, 2, 3, "Tổng:"],
        "Số công tơ": [12068236, 16719586, 9048549, ""],
        "Loại công tơ": ["3 pha", "1 pha", "1 pha", ""],
        "Địa chỉ": ["AS+ĐL FSB T12", "ĐH-Outdoor T12 (Mới)", "ĐH-Indoor CKTV (Mới)", ""],
        "CSCK": [df_fbs.iloc[i, 6] for i in range(len(df_fbs))] + [""],
        "CSĐK": [df_fbs.iloc[i, 5] for i in range(len(df_fbs))] + [""],
        "Hệ số": [1, 1, 1, ""],
        "Tổng KWh": ["", "", "", ""],
        "%": ["", "", "", ""],
        "Thanh toán (KWh)": ["", "", "", ""]
    })
    df_fbs_frame["Tổng KWh"] = [(df_fbs_frame.iloc[i, 4] - df_fbs_frame.iloc[i, 5])*df_fbs_frame.iloc[i, 6] for i in range(3)] + [""]
    df_fbs_frame_total = sum([(df_fbs_frame.iloc[i, 4] - df_fbs_frame.iloc[i, 5])*df_fbs_frame.iloc[i, 6] for i in range(3)])
    df_fbs_frame["Thanh toán (KWh)"] = [(df_fbs_frame.iloc[i, 4] - df_fbs_frame.iloc[i, 5])*df_fbs_frame.iloc[i, 6] for i in range(3)] + [df_fbs_frame_total]

    df_fbs_title = pd.DataFrame([["", "FBS"] + [""] * (df_fbs_frame.shape[1] - 2)], columns=df_fbs_frame.columns)
    df_fbs_frame = pd.concat([df_fbs_title, df_fbs_frame], ignore_index=True)

    # Bảng Tân Việt cũ
    df_tan_viet_old = df.loc[["ĐH-Outdoor T12 (Cũ)", "ĐH-Indoor T12 (CK-Cũ)", "ĐH-Indoor T12 (TTX-Cũ)"]]
    df_tan_viet_old_frame = pd.DataFrame({
        "Stt": [1, 2, 3, "Tổng:"],
        "Số công tơ": [14047859, 17736901, 14520204, ""],
        "Loại công tơ": ["3 pha", "1 pha", "1 pha", ""],
        "Địa chỉ": ["ĐH-Outdoor T12 (Cũ)", "ĐH-Indoor T12 (FSB)", "ĐH-Indoor T12 (TTX)", ""],
        "CSCK": [df_tan_viet_old.iloc[i,6] for i in range(3)] + [""],
        "CSĐK": [df_tan_viet_old.iloc[i,5] for i in range(3)] + [""],
        "Hệ số": [1, 1, 1, ""],
        "Tổng KWh": ["", "", "", ""],
        "%": ["", "", "", ""],
        "Thanh toán (KWh)": ["", "", "", ""]
    })

    df_tan_viet_old_frame['Tổng KWh'] = [(df_tan_viet_old_frame.iloc[i, 4] - df_tan_viet_old_frame.iloc[i, 5])*df_tan_viet_old_frame.iloc[i, 6] for i in range(3)] + [""]
    rate_tan_viet_old = round(df_tan_viet_old_frame.iloc[1, 7] /
                (df_tan_viet_old_frame.iloc[1, 7] + df_tan_viet_old_frame.iloc[2, 7]), 2)

    df_tan_viet_old_frame.loc[1, "%"] = rate_tan_viet_old
    df_tan_viet_old_frame_total = round(df_tan_viet_old_frame.iloc[0, 7] * rate_tan_viet_old + df_tan_viet_old_frame.iloc[1, 7], 2)
    df_tan_viet_old_frame["Thanh toán (KWh)"] = [
        df_tan_viet_old_frame.iloc[0, 7] * rate_tan_viet_old,
        df_tan_viet_old_frame.iloc[1, 7],
        "",
        df_tan_viet_old_frame_total
    ]

    df_tan_viet_old_title = pd.DataFrame([["", "Chứng khoán Tân Việt T12"] + [""] * (df_tan_viet_old_frame.shape[1] - 2)], columns=df_tan_viet_old_frame.columns)
    df_tan_viet_old_frame = pd.concat([df_tan_viet_old_title, df_tan_viet_old_frame], ignore_index=True)

    fbs_t12 = df_fbs_frame_total + df_tan_viet_old_frame_total

    # Bảng MB Hoàn Kiếm điều hòa
    df_mb_hoankiem1 = df.loc[["ĐH-Outdoor T3 (MB)", "ĐH-Outdoor T2 (MB)", "ĐH-Indoor T3+T2 (MB)", "ĐH-Outdoor T1 (MB+TTX)", "ĐH-Indoor T1 (MB)", "ĐH-Indoor T1 (TTX)"]]
    df_mb_hoankiem_frame1 = pd.DataFrame({
        "Stt": [1, 2, 3, 4, 5, 6, "Tổng:"],
        "Số công tơ": [12068239, 12063634, 12063346, 12068229, 13011012, 13011297, ""],
        "Loại công tơ": ["3 pha", "3 pha", "3 pha", "3 pha", "1 pha", "1 pha", ""],
        "Địa chỉ": [
            "ĐH tầng 3 (Outdoor)",
            "ĐH tầng 2 (Outdoor)",
            "ĐH tầng 3+2 (Indoor)",
            "ĐH tầng 1 (Outdoor)",
            "ĐH tầng 1 (Indoor)-MB",
            "ĐH tầng 1 (Indoor)-TTX",
            ""
        ],
        "CSCK": [df_mb_hoankiem1.iloc[i,6] for i in range(6)] + [""],
        "CSĐK": [df_mb_hoankiem1.iloc[i,5] for i in range(6)] + [""],
        "Hệ số": [1, 1, 1, 1, 1, 1, ""],
        "Tổng KWh": ["", "", "","", "", "", ""],
        "%": ["", "", "", "", "", "", ""],
        "Thanh toán (KWh)": ["", "", "","", "", "", ""]
    })
    df_mb_hoankiem_frame1 = pd.DataFrame(df_mb_hoankiem_frame1)
    df_mb_hoankiem_frame1["Tổng KWh"] = [(df_mb_hoankiem_frame1.iloc[i, 4] - df_mb_hoankiem_frame1.iloc[i, 5])*df_mb_hoankiem_frame1.iloc[i, 6] for i in range(6)] + [""]
    rate_mb_hoankiem1 = round(df_mb_hoankiem_frame1.iloc[4, 7]/(df_mb_hoankiem_frame1.iloc[4, 7] + df_mb_hoankiem_frame1.iloc[5, 7]), 2)

    df_mb_hoankiem_frame1["%"] = ["", "", "", rate_mb_hoankiem1, "", "", ""]
    df_mb_hoankiem_frame1_total = sum([df_mb_hoankiem_frame1.iloc[0, 7],
                                    df_mb_hoankiem_frame1.iloc[1, 7],
                                    df_mb_hoankiem_frame1.iloc[2, 7],
                                    round(df_mb_hoankiem_frame1.iloc[3, 7] * rate_mb_hoankiem1, 1),
                                    df_mb_hoankiem_frame1.iloc[4, 7]])
    df_mb_hoankiem_frame1["Thanh toán (KWh)"] = [df_mb_hoankiem_frame1.iloc[0, 7],
                                                df_mb_hoankiem_frame1.iloc[1, 7],
                                                df_mb_hoankiem_frame1.iloc[2, 7],
                                                round(df_mb_hoankiem_frame1.iloc[3, 7] * rate_mb_hoankiem1, 1),
                                                df_mb_hoankiem_frame1.iloc[4, 7],
                                                "",
                                                df_mb_hoankiem_frame1_total]
    df_mb_hoankiem_frame_title1 = pd.DataFrame([["", "MB Hoàn Kiếm 1"] + [""] * (df_mb_hoankiem_frame1.shape[1] - 2)], columns=df_mb_hoankiem_frame1.columns)
    df_mb_hoankiem_frame1 = pd.concat([df_mb_hoankiem_frame_title1, df_mb_hoankiem_frame1], ignore_index=True)

    # Bảng MB Hoàn Kiếm ánh sáng
    df_mb_hoankiem2 = df.loc[["AS + ĐL T3+T2+T1 (MB)"]]
    df_mb_hoankiem_frame2 = pd.DataFrame({
        "Stt": [1, "Tổng:"],
        "Số công tơ": [403189, ""],
        "Loại công tơ": ["3 pha", ""],
        "Địa chỉ": ["AS + ĐL tầng 1,2,3", ""],
        "CSCK": [df_mb_hoankiem2.iloc[0,6], ""],
        "CSĐK": [df_mb_hoankiem2.iloc[0,5], ""],
        "Hệ số": [1, ""],
        "Tổng KWh": ["",""],
        "%": ["", ""],
        "Thanh toán (KWh)": ["",""]
    })

    df_mb_hoankiem_frame2["Tổng KWh"] = [(df_mb_hoankiem_frame2.iloc[i, 4] - df_mb_hoankiem_frame2.iloc[i, 5])*df_mb_hoankiem_frame2.iloc[i, 6] for i in range(1)] + [""]
    df_mb_hoankiem_frame2_total = sum([(df_mb_hoankiem_frame2.iloc[i, 4] - df_mb_hoankiem_frame2.iloc[i, 5])*df_mb_hoankiem_frame2.iloc[i, 6] for i in range(1)])
    df_mb_hoankiem_frame2["Thanh toán (KWh)"] = [df_mb_hoankiem_frame2_total, df_mb_hoankiem_frame2_total]
    df_mb_hoankiem_frame2_title = pd.DataFrame([["", "MB Hoàn Kiếm 2"] + [""] * (df_mb_hoankiem_frame2.shape[1] - 2)], columns=df_mb_hoankiem_frame2.columns)
    df_mb_hoankiem_frame2 = pd.concat([df_mb_hoankiem_frame2_title, df_mb_hoankiem_frame2], ignore_index=True)

    df_mb_hoankiem_frame_total = df_mb_hoankiem_frame1_total + df_mb_hoankiem_frame2_total

    # Bảng Giovani điều hòa
    df_giovani1 = df.loc[["Outdoor 1-T1", "Outdoor 2-T1", "Indoor-GIOVANI", "Indoor- TTX"]]
    df_giovani_frame1 = pd.DataFrame({
        "Stt": [1, 2, 3, 4, "Tổng:"],
        "Số công tơ": [14038150, 14038145, 15009568, 15012663, ""],
        "Loại công tơ": ["3 pha", "3 pha", "1 pha", "1 pha", ""],
        "Địa chỉ": [
            "ĐH-Outdoor M1",
            "ĐH-Outdoor M2",
            "ĐH-Indoor N1 (GVN)",
            "ĐH-Indoor N2 (TTX)",
            ""
        ],
        "CSCK": [df_giovani1.iloc[i,6] for i in range(4)] + [""],
        "CSĐK": [df_giovani1.iloc[i,5] for i in range(4)] + [""],
        "Hệ số": [1, 1, 1, 1, ""],
        "Tổng KWh": ["", "", "", "", ""],
        "%": ["", "", "", "", ""],
        "Thanh toán (KWh)": ["", "", "", "", ""]
    })
    df_giovani_frame1['Tổng KWh'] = [(df_giovani_frame1.iloc[i, 4] - df_giovani_frame1.iloc[i, 5])*df_giovani_frame1.iloc[i, 6] for i in range(4)] + [""]
    df_giovani_frame1.loc[2,"%"] = round(df_giovani_frame1.iloc[2, 7] / (df_giovani_frame1.iloc[2, 7] + df_giovani_frame1.iloc[3, 7]), 2)
    df_giovani_frame1.loc[0, "Thanh toán (KWh)"] = (df_giovani_frame1.iloc[0, 7] + df_giovani_frame1.iloc[1, 7]) * df_giovani_frame1.loc[2, "%"]
    df_giovani_frame1.loc[2, "Thanh toán (KWh)"] = df_giovani_frame1.iloc[2, 7]
    df_giovani_frame1_total = df_giovani_frame1.loc[0, "Thanh toán (KWh)"] + df_giovani_frame1.loc[2, "Thanh toán (KWh)"]
    df_giovani_frame1.loc[4, "Thanh toán (KWh)"] = df_giovani_frame1_total
    df_giovani_frame1_title = pd.DataFrame([["", "Giovani1"] + [""] * (df_giovani_frame1.shape[1] - 2)], columns=df_giovani_frame1.columns)
    df_giovani_frame1 = pd.concat([df_giovani_frame1_title, df_giovani_frame1], ignore_index=True)

    # Bảng Giovani ánh sáng
    df_giovani2 = df.loc[["AS+ĐL - GIOVANI"]]
    df_giovani_frame2 = pd.DataFrame({
        "Stt": [1, "Tổng:"],
        "Số công tơ": [10511904, ""],
        "Loại công tơ": ["3 pha", ""],
        "Địa chỉ": ["AS + ĐL DVT1", ""],
        "CSCK": [df_giovani2.iloc[0, 6], ""],
        "CSĐK": [df_giovani2.iloc[0, 5], ""],
        "Hệ số": [1, ""],
        "Tổng KWh": ["", ""],
        "%": ["", ""],
        "Thanh toán (KWh)": ["", ""]
    })
    df_giovani_frame2["Tổng KWh"] = [(df_giovani_frame2.iloc[i, 4] - df_giovani_frame2.iloc[i, 5])*df_giovani_frame2.iloc[i, 6] for i in range(1)] + [""]
    df_giovani_frame2_total = sum([(df_giovani_frame2.iloc[i, 4] - df_giovani_frame2.iloc[i, 5])*df_giovani_frame2.iloc[i, 6] for i in range(1)])
    df_giovani_frame2["Thanh toán (KWh)"] = [df_giovani_frame2_total, df_giovani_frame2_total]
    df_giovani_frame2_title = pd.DataFrame([["", "Giovani2"] + [""] * (df_giovani_frame2.shape[1] - 2)], columns=df_giovani_frame2.columns)
    df_giovani_frame2 = pd.concat([df_giovani_frame2_title, df_giovani_frame2], ignore_index=True)

    df_giovani_frame_total = df_giovani_frame1_total + df_giovani_frame2_total

    # Bảng dịch vụ công ty GME
    df_gme = df.loc[["AS + ĐL T4 (GME)", "ĐH-Outdoor T4 (GME +TTX)", "ĐH-Indoor T4 (GME)", "ĐH-Indoor T4 (TTX)"]]
    df_gme_frame = pd.DataFrame({
        "Stt": [4, 1, 3, 2, "Tổng:"],
        "Số công tơ": [16698180, 16705013, 16702810, 16702656, ""],
        "Loại công tơ": ["1 pha", "3 pha", "1 pha", "1 pha", ""],
        "Địa chỉ": [
            "AS + ĐL T4 (GME)",
            "ĐH-Outdoor T4",
            "ĐH-Indoor T4 (GME)",
            "ĐH-Indoor T4 (TTX)",
            ""
        ],
        "CSCK": [df_gme.iloc[i, 6] for i in range(df_gme.shape[0])] + [""],
        "CSĐK": [df_gme.iloc[i, 5] for i in range(df_gme.shape[0])] + [""],
        "Hệ số": [1, 1, 1, 1, ""],
        "Tổng KWh": ["", "", "", "", ""],
        "%": ["", "", "", "", ""],
        "Thanh toán (KWh)": ["", "", "", "", ""]
    })
    df_gme_frame['Tổng KWh'] = [(df_gme_frame.iloc[i, 4] - df_gme_frame.iloc[i, 5])*df_gme_frame.iloc[i, 6] for i in range(df_gme_frame.shape[0]-1)] + [""]
    df_gme_frame.loc[2, "%"] = round(df_gme_frame.iloc[2, 7] / (df_gme_frame.iloc[2, 7] + df_gme_frame.iloc[3, 7]), 3)
    df_gme_frame.loc[0, "Thanh toán (KWh)"] = df_gme_frame.iloc[0, 7]
    df_gme_frame.loc[1, "Thanh toán (KWh)"] = round(df_gme_frame.iloc[1, 7]*df_gme_frame.loc[2, "%"], 1)
    df_gme_frame.loc[2, "Thanh toán (KWh)"] = df_gme_frame.iloc[2, 7]
    df_gme_frame_total = sum([df_gme_frame.loc[i, "Thanh toán (KWh)"] for i in range(3)])
    df_gme_frame.loc[4, "Thanh toán (KWh)"] = df_gme_frame_total
    df_gme_frame_title = pd.DataFrame([["", "GME"] + [""] * (df_gme_frame.shape[1] - 2)], columns=df_gme_frame.columns)
    df_gme_frame = pd.concat([df_gme_frame_title, df_gme_frame], ignore_index=True)

    total_service = round(df_fbs_frame_total + df_tan_viet_old_frame_total + df_mb_hoankiem_frame1_total + df_mb_hoankiem_frame2_total + df_giovani_frame1_total + df_giovani_frame2_total + df_gme_frame_total, 1)

    # Bảng màn hình led 5LTK
    df_led = df.loc[["Màn hình Led-5LTK"]]
    df_led_frame = pd.DataFrame({
        "Stt": [1, "Tổng"],
        "Số công tơ": [18754117, ""],
        "Loại công tơ": ["3 pha", ""],
        "Địa chỉ": ["Màn hình Led hè 5 LTK", ""],
        "CSCK": [df_led.iloc[0, 6]] + [""],
        "CSĐK": [df_led.iloc[0, 5]] + [""],
        "Hệ số": [20, ""],
        "Tổng KWh": ["", ""],
        "%": ["", ""],
        "Thanh toán (KWh)": ["", ""]
    })
    df_led_frame["Tổng KWh"] = [(df_led_frame.iloc[i, 4] - df_led_frame.iloc[i, 5])*df_led_frame.iloc[i, 6] for i in range(1)] + [""]
    df_led_frame.loc[0, "Thanh toán (KWh)"] = df_led_frame.iloc[0, 7]
    df_led_frame_total = sum([df_led_frame.iloc[i, 7] for i in range(1)])
    df_led_frame["Thanh toán (KWh)"] = [df_led_frame_total, df_led_frame_total]
    df_led_frame_title = pd.DataFrame([["", "Màn hình Led"] + [""] * (df_led_frame.shape[1] - 2)], columns=df_led_frame.columns)
    df_led_frame = pd.concat([df_led_frame_title, df_led_frame], ignore_index=True)

    # Bảng hội trường tầng 10
    df_hoitruong_t10 = df.loc[["T11-O3 (21)", "T11-O4 (22)", "T11-O5 (23)", "T11-O6 (24)", "T11-Hội Trường"]]
    df_hoitruong_t10_frame = pd.DataFrame({
        "Stt": [1, 2, 3, 4, 5, "Tổng"],
        "Số công tơ": [16737292, 16737284, 16737323, 16737346, 15017041, ""],
        "Loại công tơ": ["3 pha"] * 5 + [""],
        "Địa chỉ": ["T11-O3 (21)", "T11-O4 (22)", "T11-O5 (23)", "T11-O6 (24)", "T11-Hội Trường", ""],
        "CSCK": [df_hoitruong_t10.iloc[i, 6] for i in range(df_hoitruong_t10.shape[0])] + [""],
        "CSĐK": [df_hoitruong_t10.iloc[i, 5] for i in range(df_hoitruong_t10.shape[0])] + [""],
        "Hệ số": [1, 1, 1, 1, 1, ""],
        "Tổng KWh": ["", "", "", "", "", ""],
        "%": ["", "", "", "", "", ""],
        "Thanh toán (KWh)": ["", "", "", "", "", ""]
    })
    df_hoitruong_t10_frame["Tổng KWh"] = [(df_hoitruong_t10_frame.iloc[i, 4] - df_hoitruong_t10_frame.iloc[i, 5])*df_hoitruong_t10_frame.iloc[i, 6] for i in range(df_hoitruong_t10.shape[0])] + [""]
    df_hoitruong_t10_frame.loc[0, "Thanh toán (KWh)"] = df_hoitruong_t10_frame.iloc[0, 7]
    df_hoitruong_t10_frame_total = sum([df_hoitruong_t10_frame.iloc[i, 7] for i in range(5)])
    df_hoitruong_t10_frame.loc[5, "Thanh toán (KWh)"] = df_hoitruong_t10_frame_total
    df_hoitruong_t10_frame_title = pd.DataFrame([["", "Hội trườngT10"] + [""] * (df_hoitruong_t10_frame.shape[1] - 2)], columns=df_hoitruong_t10_frame.columns)
    df_hoitruong_t10_frame = pd.concat([df_hoitruong_t10_frame_title, df_hoitruong_t10_frame], ignore_index=True)

    # Bảng sảnh tầng 11
    df_t11 = df.loc[["T11-O1 (19)", "T11-O2 (20)", "T11-Sảnh"]]
    df_t11_frame = pd.DataFrame({
        "Stt": [1, 2, 3, "Tổng"],
        "Số công tơ": [16736345, 16737344, 15017087, ""],
        "Loại công tơ": ["3 pha", "3 pha", "3 pha", ""],
        "Địa chỉ": ["T11-O1 (19)", "T11-O2 (20)", "T11-indoor Sảnh", ""],
        "CSCK": [df_t11.iloc[i, 6] for i in range(3)] + [""],
        "CSĐK": [df_t11.iloc[i, 5] for i in range(3)] + [""],
        "Hệ số": [1, 1, 1, ""],
        "Tổng KWh": ["", "", "", ""],
        "%": ["", "", "", ""],
        "Thanh toán (KWh)": ["", "", "", ""]
    })
    df_t11_frame['Tổng KWh'] = [(df_t11_frame.iloc[i, 4] - df_t11_frame.iloc[i, 5])*df_t11_frame.iloc[i, 6] for i in range(3)] + [""]
    df_t11_frame.loc[0, "Thanh toán (KWh)"] = sum([df_t11_frame.iloc[i, 7] for i in range(3)])
    df_t11_frame_total = sum([df_t11_frame.iloc[i, 7] for i in range(3)])
    df_t11_frame.loc[3, "Thanh toán (KWh)"] = df_t11_frame_total
    df_t11_frame_title = pd.DataFrame([["", "Sảnh T11"] + [""] * (df_t11_frame.shape[1] - 2)], columns=df_t11_frame.columns)
    df_t11_frame = pd.concat([df_t11_frame_title, df_t11_frame], ignore_index=True)

    # Bảng sảnh tầng 10
    df_sanh_t10 = df.loc[["T10-O1 (25)", "T10-O2 (26)", "T10-Truyền Thống", "T10-Sảnh"]]
    df_sanh_t10_frame = pd.DataFrame({
        "Stt": [1, 2, 3, 4, "Tổng"],
        "Số công tơ": [16737283, 16737297, 15017132, 15017120, ""],
        "Loại công tơ": ["3 pha"] * 4 + [""],
        "Địa chỉ": ["T10-O1 (25)", "T10-O2 (26)", "T10-Truyền Thống", "T10-Sảnh", ""],
        "CSCK": [df_sanh_t10.iloc[i, 6] for i in range(df_sanh_t10.shape[0])] + [""],
        "CSĐK": [df_sanh_t10.iloc[i, 5] for i in range(df_sanh_t10.shape[0])] + [""],
        "Hệ số": [1, 1, 1, 1, ""],
        "Tổng KWh": ["", "", "", "", ""],
        "%": ["", "", "", "", ""],
        "Thanh toán (KWh)": ["", "", "", "", ""]
    })
    df_sanh_t10_frame['Tổng KWh'] = [(df_sanh_t10_frame.iloc[i, 4] - df_sanh_t10_frame.iloc[i, 5])*df_sanh_t10_frame.iloc[i, 6] for i in range(df_sanh_t10.shape[0])] + [""]
    df_sanh_t10_frame.loc[3, '%'] = round(df_sanh_t10_frame.iloc[3, 7] / (df_sanh_t10_frame.iloc[3, 7] + df_sanh_t10_frame.iloc[2, 7]), 2)
    df_sanh_t10_frame.loc[1, "Thanh toán (KWh)"] = (df_sanh_t10_frame.iloc[0, 7] + df_sanh_t10_frame.iloc[1, 7])*df_sanh_t10_frame.loc[3, '%']
    df_sanh_t10_frame.loc[3, "Thanh toán (KWh)"] = df_sanh_t10_frame.iloc[3, 7]
    df_sanh_t10_frame_total = df_sanh_t10_frame.loc[1, "Thanh toán (KWh)"] + df_sanh_t10_frame.loc[3, "Thanh toán (KWh)"]
    df_sanh_t10_frame.loc[4, "Thanh toán (KWh)"] = df_sanh_t10_frame_total
    df_sanh_t10_frame_title = pd.DataFrame([["", "Sảnh T10"] + [""] * (df_sanh_t10_frame.shape[1] - 2)], columns=df_sanh_t10_frame.columns)
    df_sanh_t10_frame = pd.concat([df_sanh_t10_frame_title, df_sanh_t10_frame], ignore_index=True)

    # Bảng phòng truyền thống tầng 10
    df_phong_truyen_thong_t10 = df.loc[["T10-O1 (25)", "T10-O2 (26)", "T10-Truyền Thống", "T10-Sảnh"]]
    df_phong_truyen_thong_t10_frame= pd.DataFrame({
        "Stt": [1, 2, 3, 4, "Tổng"],
        "Số công tơ": [16737283, 16737297, 15017132, 15017120, ""],
        "Loại công tơ": ["3 pha"] * 4 + [""],
        "Địa chỉ": ["T10-O1 (25)", "T10-O2 (26)", "T10-Truyền Thống", "T10-Sảnh", ""],
        "CSCK": [df_phong_truyen_thong_t10.iloc[i, 6] for i in range(df_phong_truyen_thong_t10.shape[0])] + [""],
        "CSĐK": [df_phong_truyen_thong_t10.iloc[i, 5] for i in range(df_phong_truyen_thong_t10.shape[0])] + [""],
        "Hệ số": [1, 1, 1, 1, ""],
        "Tổng KWh": ["", "", "", "", ""],
        "%": ["", "", "", "", ""],
        "Thanh toán (KWh)": ["", "", "", "", ""]
    })
    df_phong_truyen_thong_t10_frame["Tổng KWh"] = [(df_phong_truyen_thong_t10_frame.iloc[i, 4] - df_phong_truyen_thong_t10_frame.iloc[i, 5])*df_phong_truyen_thong_t10_frame.iloc[i, 6]
                                                for i in range(df_phong_truyen_thong_t10.shape[0])] + [""]
    df_phong_truyen_thong_t10_frame.loc[2, "%"] = round(df_phong_truyen_thong_t10_frame.iloc[2, 7] / (df_phong_truyen_thong_t10_frame.iloc[2, 7] + df_phong_truyen_thong_t10_frame.iloc[3, 7]), 2)
    df_phong_truyen_thong_t10_frame.loc[0, "Thanh toán (KWh)"] = (df_phong_truyen_thong_t10_frame.iloc[0, 7] + df_phong_truyen_thong_t10_frame.iloc[1, 7])*df_phong_truyen_thong_t10_frame.loc[2, "%"]
    df_phong_truyen_thong_t10_frame.loc[2, "Thanh toán (KWh)"] = df_phong_truyen_thong_t10_frame.iloc[2, 7]
    df_phong_truyen_thong_t10_frame_total = df_phong_truyen_thong_t10_frame.loc[0, "Thanh toán (KWh)"] + df_phong_truyen_thong_t10_frame.loc[2, "Thanh toán (KWh)"]
    df_phong_truyen_thong_t10_frame.loc[4, "Thanh toán (KWh)"] = df_phong_truyen_thong_t10_frame_total
    df_phong_truyen_thong_t10_frame_title = pd.DataFrame([["", "Phòng Truyền Thống T10"] + [""] * (df_phong_truyen_thong_t10_frame.shape[1] -2)], columns=df_phong_truyen_thong_t10_frame.columns)
    df_phong_truyen_thong_t10_frame = pd.concat([df_phong_truyen_thong_t10_frame_title, df_phong_truyen_thong_t10_frame], ignore_index=True)

    # Bảng cafe báo chí tầng 9
    df_cafe_t9 = df.loc[["T9-OC&AS Dịch vụ", "T9-OC khu bếp", "T9-O1 (27)", "T9-O2 (28)", "T9-DH Dịch vụ"]]
    df_cafe_t9
    df_cafe_t9_frame = pd.DataFrame({
        "Stt": [1, 2, 3, 4, 5, "Tổng"],
        "Số công tơ": [16719505, 16719580, 16737351, 16737350, 15016996, ""],
        "Loại công tơ": ["3 pha"] * 5 + [""],
        "Địa chỉ": [
            "AS+ĐL Dịch vụ tầng 9",
            "T9-OC khu bếp",
            "T9-O1 (27)",
            "T9-O2 (28)",
            "T9-DH Dịch vụ",
            ""
        ],
        "CSCK": [df_cafe_t9.iloc[i, 6] for i in range(df_cafe_t9.shape[0])] + [""],
        "CSĐK": [df_cafe_t9.iloc[i, 5] for i in range(df_cafe_t9.shape[0])] + [""],
        "Hệ số": [1, 1, 1, 1, 1, ""],
        "Tổng KWh": ["", "", "", "", "", ""],
        "%": ["", "", "", "", "", ""],
        "Thanh toán (KWh)": ["", "", "", "", "", ""]
    })
    df_cafe_t9_frame['Tổng KWh'] = [(df_cafe_t9_frame.iloc[i, 4] - df_cafe_t9_frame.iloc[i, 5])*df_cafe_t9_frame.iloc[i, 6] for i in range(df_cafe_t9.shape[0])] + [""]
    df_cafe_t9_frame['Thanh toán (KWh)'] = df_cafe_t9_frame['Tổng KWh']
    df_cafe_t9_frame_total = sum([df_cafe_t9_frame.iloc[i, 7] for i in range(5)])
    df_cafe_t9_frame.loc[5, "Thanh toán (KWh)"] = df_cafe_t9_frame_total
    df_cafe_t9_frame_title = pd.DataFrame([["", "Cafe T9"] + [""] * (df_cafe_t9_frame.shape[1] - 2)], columns=df_cafe_t9_frame.columns)
    df_cafe_t9_frame = pd.concat([df_cafe_t9_frame_title, df_cafe_t9_frame], ignore_index=True)

    # Bảng nhà ăn tầng 9
    df_nha_an_t9 = df.loc[["T9-OC&AS Nhà ăn", "T9-O3 (48)", "T9-O4 (49)", "T9-O5 (50)", "T9-DH Nhà ăn"]]
    df_nha_an_t9_frame = pd.DataFrame({
        "Stt": [1, 2, 3, 4, 5, "Tổng"],
        "Số công tơ": [16711486, 16737293, 16736346, 16737286, 16655845, ""],
        "Loại công tơ": ["3 pha"] * 5 + [""],
        "Địa chỉ": [
            "AS+ĐL Nhà ăn tầng 9",
            "T9-O3 (48)",
            "T9-O4 (49)",
            "T9-O5 (50)",
            "T9-DH Nhà ăn",
            ""
        ],
        "CSCK": [df_nha_an_t9.iloc[i, 6] for i in range(df_nha_an_t9.shape[0])] + [""],
        "CSĐK": [df_nha_an_t9.iloc[i, 5] for i in range(df_nha_an_t9.shape[0])] + [""],
        "Hệ số": [1, 1, 1, 1, 1, ""],
        "Tổng KWh": ["", "", "", "", "", ""],
        "%": ["", "", "", "", "", ""],
        "Thanh toán (KWh)": ["", "", "", "", "", ""]
    })
    df_nha_an_t9_frame['Tổng KWh'] = [(df_nha_an_t9_frame.iloc[i, 4] - df_nha_an_t9_frame.iloc[i, 5])*df_nha_an_t9_frame.iloc[i, 6] for i in range(df_nha_an_t9.shape[0])] + [""]
    df_nha_an_t9_frame['Thanh toán (KWh)'] = df_nha_an_t9_frame['Tổng KWh']
    df_nha_an_t9_frame_total = sum([df_nha_an_t9_frame.iloc[i, 7] for i in range(5)])
    df_nha_an_t9_frame.loc[5, "Thanh toán (KWh)"] = df_nha_an_t9_frame_total
    df_nha_an_t9_frame_title = pd.DataFrame([["", "Nhà ăn T9"] + [""] * (df_nha_an_t9_frame.shape[1] - 2)], columns=df_nha_an_t9_frame.columns)
    df_nha_an_t9_frame = pd.concat([df_nha_an_t9_frame_title, df_nha_an_t9_frame], ignore_index=True)

    # Bảng Trung tâm hợp tác Quốc tế Thông tấn phần văn phòng
    df_vna8_vp = df.loc[["AS+ĐL 8THĐ", "ĐH 8THĐ", "33LTT",
                    "AS18THĐ", "ĐH18THĐ", "AS20THĐ", "ĐH20THĐ",
                    "AS-T6-11THĐ", "ĐH-T6-11THĐ", "AS-P604-11THĐ", "ĐH-P604-11THĐ", "AS T5.11THĐ", "ĐH T5.11THĐ",
                    "ĐH-T7 P701-703-705", "AS-T7 P701-703-705", "ĐH-T7 P702-704-706", "AS-T7 P702-704-706", "ĐH-T7 P707", "AS-T7 P707", "ĐH-T7 P708", "AS-T7 P708"]]
    df_vna8_vp_frame = pd.DataFrame({
        "Stt": [1, "", "", "", 2, "2.1", "", "", "", "", "2.2", "", "", "", "", "", "", "2.3", "", "", "", "", "", "", "", "", 3, "Tổng"],
        "Số công tơ": ["", 14308090, 14308130, 372755, "CT mới", "", 15485, 15485, 15485, 15485, "", "", "", "", "", "", "", "", 30094, 37553, 30098, 30138, 13579, 52583, 13464, 277745, 99178786, ""],
        "Loại công tơ": ["", "3 pha", "3 pha", "3 pha", "3 pha", "", "3 pha", "3 pha", "3 pha", "3 pha", "", "3 pha", "3 pha", "1 pha", "1 pha", "3 pha", "3 pha", "", "3 pha", "3 pha", "3 pha", "3 pha", "1 pha", "1 pha", "1 pha", "1 pha", "3 pha", ""],
        "Địa chỉ": ["Khu vực số 8THĐ", "AS 8THĐ (từ 5LTK)", "ĐH 8THĐ (từ 5LTK)", "Số 33LTT (từ 8THĐ)", "BA - 630KVA", "18-20 THĐ", "AS18. THĐ", "ĐH18. THĐ", "AS20. THĐ", "ĐH20. THĐ",
                    "T6.11.THĐ Sử dụng", "AS T6.11THĐ", "ĐH T6.11THĐ", "AS P604.11THĐ", "ĐH P604.11THĐ", "AS T5.11THĐ", "ĐH T5.11THĐ", "T7.11.THĐ Sử dụng", "ĐH.P701-3-5", "AS.P701-3-5",
                    "ĐH.P702-4-6", "AS.P702-4-6", "ĐH. P707", "AS. P707", "ĐH. P708", "AS. P708", "Nguồn của Sphon", ""],
        "CSCK": ["", df_vna8_vp.iloc[0, 6], df_vna8_vp.iloc[1, 6], df_vna8_vp.iloc[2, 6], 0, "",
                df_vna8_vp.iloc[3, 6], df_vna8_vp.iloc[4, 6], df_vna8_vp.iloc[5, 6], df_vna8_vp.iloc[6, 6], "",
                df_vna8_vp.iloc[7, 6], df_vna8_vp.iloc[8, 6], df_vna8_vp.iloc[9, 6], df_vna8_vp.iloc[10, 6], df_vna8_vp.iloc[11, 6], df_vna8_vp.iloc[12, 6], "",
                df_vna8_vp.iloc[13, 6], df_vna8_vp.iloc[14, 6], df_vna8_vp.iloc[15, 6], df_vna8_vp.iloc[16, 6], df_vna8_vp.iloc[17, 6], df_vna8_vp.iloc[18, 6], df_vna8_vp.iloc[19, 6], df_vna8_vp.iloc[20, 6], "", ""],
        "CSĐK": ["", df_vna8_vp.iloc[0, 5], df_vna8_vp.iloc[1, 5], df_vna8_vp.iloc[2, 5], 0, "",
                df_vna8_vp.iloc[3, 5], df_vna8_vp.iloc[4, 5], df_vna8_vp.iloc[5, 5], df_vna8_vp.iloc[6, 5], "",
                df_vna8_vp.iloc[7, 5], df_vna8_vp.iloc[8, 5], df_vna8_vp.iloc[9, 5], df_vna8_vp.iloc[10, 5], df_vna8_vp.iloc[11, 5], df_vna8_vp.iloc[12, 5], "",
                df_vna8_vp.iloc[13, 5], df_vna8_vp.iloc[14, 5], df_vna8_vp.iloc[15, 5], df_vna8_vp.iloc[16, 5], df_vna8_vp.iloc[17, 5], df_vna8_vp.iloc[18, 5], df_vna8_vp.iloc[19, 5], df_vna8_vp.iloc[20, 5], "", ""],
        "Hệ số": ["", 60, 60, 40, 200, "", 2, 2, 2, 2, "", 10, 20, 1, 1, 10, 20, "", 20, 10, 20, 10, 1, 1, 1, 1, 20, ""],
        "Tổng KWh": [""] * 28,
        "%": [""] * 28,
        "Thanh toán (KWh)": [""] * 28
    })
    df_vna8_vp_frame['Tổng KWh'] = [""] + [(df_vna8_vp_frame.iloc[i, 4] - df_vna8_vp_frame.iloc[i, 5]) * df_vna8_vp_frame.iloc[i, 6] for i in range(1, 4)] + [200] +\
                                    [""] + [(df_vna8_vp_frame.iloc[i, 4] - df_vna8_vp_frame.iloc[i, 5]) * df_vna8_vp_frame.iloc[i, 6] for i in range(6, 10)] +\
                                    [""] + [(df_vna8_vp_frame.iloc[i, 4] - df_vna8_vp_frame.iloc[i, 5]) * df_vna8_vp_frame.iloc[i, 6] for i in range(11, 17)] +\
                                    [""] + [(df_vna8_vp_frame.iloc[i, 4] - df_vna8_vp_frame.iloc[i, 5]) * df_vna8_vp_frame.iloc[i, 6] for i in range(18, 26)] + [20] + [""]

    df_vna8_vp_frame.loc[1, "Thanh toán (KWh)"] =  sum([(df_vna8_vp_frame.iloc[i, 4] - df_vna8_vp_frame.iloc[i, 5]) * df_vna8_vp_frame.iloc[i, 6] for i in range(1, 4)])
    df_vna8_vp_frame.loc[6, "Thanh toán (KWh)"] =  sum([(df_vna8_vp_frame.iloc[i, 4] - df_vna8_vp_frame.iloc[i, 5]) * df_vna8_vp_frame.iloc[i, 6] for i in range(6, 10)])
    df_vna8_vp_frame.loc[11, "Thanh toán (KWh)"] = sum([(df_vna8_vp_frame.iloc[i, 4] - df_vna8_vp_frame.iloc[i, 5]) * df_vna8_vp_frame.iloc[i, 6] for i in range(11, 17)])
    df_vna8_vp_frame.loc[18, "Thanh toán (KWh)"] = sum([(df_vna8_vp_frame.iloc[i, 4] - df_vna8_vp_frame.iloc[i, 5]) * df_vna8_vp_frame.iloc[i, 6] for i in range(18, 26)])
    df_vna8_vp_frame_total = sum([df_vna8_vp_frame.loc[1, "Thanh toán (KWh)"], df_vna8_vp_frame.loc[6, "Thanh toán (KWh)"], df_vna8_vp_frame.loc[11, "Thanh toán (KWh)"], df_vna8_vp_frame.loc[18, "Thanh toán (KWh)"]])
    df_vna8_vp_frame.loc[27, "Thanh toán (KWh)"] = df_vna8_vp_frame_total
    df_vna8_vp_frame_title = pd.DataFrame([["", "VNA8-VP"] + [""] * (df_vna8_vp_frame.shape[1] - 2)], columns=df_vna8_vp_frame.columns)
    df_vna8_vp_frame = pd.concat([df_vna8_vp_frame_title, df_vna8_vp_frame], ignore_index=True)


    # Bảng Trung tâm hợp tác Quốc tế Thông tấn phần PG Bank
    df_vna8_pgb = df.loc[["ĐH  -T1 (11THĐ)", "AS - T1 (11THĐ)", "ĐH  -T2 (11THĐ)", "AS - T2 (11THĐ)"]]
    df_vna8_pgb_frame = pd.DataFrame({
        "Stt": [1, 2, 3, 4, "Tổng"],
        "Số công tơ": [12067352, 13012746, 13010295, 13010298, ""],
        "Loại công tơ": ["1 pha", "1 pha", "1 pha", "1 pha", ""],
        "Địa chỉ": ["ĐH  -T1 (11THĐ)", "AS - T1 (11THĐ)", "ĐH  -T2 (11THĐ)", "AS - T2 (11THĐ)", ""],
        "CSCK": [df_vna8_pgb.iloc[i, 6] for i in range(df_vna8_pgb.shape[0])] + [""],
        "CSĐK": [df_vna8_pgb.iloc[i, 5] for i in range(df_vna8_pgb.shape[0])] + [""],
        "Hệ số": [1, 1, 1, 1, ""],
        "Tổng KWh": ["", "", "", "", ""],
        "%": ["", "", "", "", ""],
        "Thanh toán (KWh)": ["", "", "", "", ""]
    })
    df_vna8_pgb_frame["Tổng KWh"] = [(df_vna8_pgb_frame.iloc[i, 4] - df_vna8_pgb_frame.iloc[i, 5]) * df_vna8_pgb_frame.iloc[i, 6] for i in range(df_vna8_pgb.shape[0])] + [""]
    df_vna8_pgb_frame["Thanh toán (KWh)"] = df_vna8_pgb_frame["Tổng KWh"]
    df_vna8_pgb_frame_total = sum([df_vna8_pgb_frame.iloc[i, 7] for i in range(4)])
    df_vna8_pgb_frame.loc[4, "Thanh toán (KWh)"] = df_vna8_pgb_frame_total
    df_vna8_pgb_frame_title = pd.DataFrame([["", "VNA8-PGB"] + [""] * (df_vna8_pgb_frame.shape[1] - 2)], columns=df_vna8_pgb_frame.columns)
    df_vna8_pgb_frame = pd.concat([df_vna8_pgb_frame_title, df_vna8_pgb_frame], ignore_index=True)
    df_vna8_total = df_vna8_vp_frame_total + df_vna8_pgb_frame_total

    # Bảng Vietnam News
    df_vnnews = df.loc[["OC+AS News T9+10", "UPS News T9+10", "ĐH-Indoor T10 (News)", "ĐH-Indoor T10 (XB- A Sơn)", "ĐH-Outdoor T10 (Mới)", "ĐH-Outdoor T9 ", "ĐH-Indoor T9 (News)"]]
    df_vnnews_frame = pd.DataFrame({
        "Stt": [1, 2, 6, 7, 3, 4, 5, "Tổng"],
        "Số công tơ": [18754273, 18791956, 18736254, 18781404, 18791952, 16719584, 18736247, ""],
        "Loại công tơ": ["3 pha", "3 pha", "3 pha", "1 pha", "3 pha", "3 pha", "3 pha", ""],
        "Địa chỉ": [
            "OC+AS News T9+10", "UPS News T9+10", "ĐH-Indoor T10 (News)", "ĐH-Indoor T10 (XB)",
            "ĐH-Outdoor T10 (Mới)", "ĐH-Outdoor T9", "ĐH-Indoor T9 (News)", ""
        ],
        "CSCK": [df_vnnews.iloc[i, 6] for i in range(df_vnnews.shape[0])] + [""],
        "CSĐK": [df_vnnews.iloc[i, 5] for i in range(df_vnnews.shape[0])] + [""],
        "Hệ số": [20, 1, 1, 1, 1, 1, 1, ""],
        "Tổng KWh": ["", "", "", "", "", "", "", ""],
        "%": ["", "", "", "", "", "", "", ""],
        "Thanh toán (KWh)": ["", "", "", "", "", "", "", ""]
    })
    df_vnnews_frame['Tổng KWh'] = [(df_vnnews_frame.iloc[i, 4] - df_vnnews_frame.iloc[i, 5]) * df_vnnews_frame.iloc[i, 6] for i in range(df_vnnews.shape[0])] + [""]
    df_vnnews_frame.loc[2, "%"] = round(df_vnnews_frame.iloc[2, 7] / (df_vnnews_frame.iloc[2, 7] + df_vnnews_frame.iloc[3, 7]), 2)
    df_vnnews_frame['Thanh toán (KWh)'] = df_vnnews_frame['Tổng KWh']
    df_vnnews_frame.loc[3, "Thanh toán (KWh)"] = 0
    df_vnnews_frame.loc[4, "Thanh toán (KWh)"] = df_vnnews_frame.iloc[4, 7] * df_vnnews_frame.loc[2, "%"]
    df_vnnews_frame_total = round(sum([df_vnnews_frame.iloc[i, 9] for i in range(7)]), 1)
    df_vnnews_frame.loc[7, "Thanh toán (KWh)"] = df_vnnews_frame_total
    df_vnnews_frame_title = pd.DataFrame([["", "VN NEWS"] + [""] * (df_vnnews_frame.shape[1] - 2)], columns=df_vnnews_frame.columns)
    df_vnnews_frame = pd.concat([df_vnnews_frame_title, df_vnnews_frame], ignore_index=True)

    # Tổng hợp sheet dịch vụ
    dich_vu_sheet_frame = pd.concat([df_fbs_frame, df_tan_viet_old_frame, df_mb_hoankiem_frame1, df_mb_hoankiem_frame2,
                            df_giovani_frame1, df_giovani_frame2, df_gme_frame, df_led_frame,
                            df_hoitruong_t10_frame, df_t11_frame, df_sanh_t10_frame, df_phong_truyen_thong_t10_frame,
                            df_cafe_t9_frame, df_nha_an_t9_frame, df_vna8_vp_frame, df_vna8_pgb_frame, df_vnnews_frame], ignore_index=True)
    # II. Phần tổng hợp
    df_tonghop = df.loc[["PXHN", "Bảo Việt mới", "TBA1-5LTK 800KVA-22/0.4KV(AS,OC)", "TBA2-5LTK 800KVA-22/0.4KV(ĐH)", "TBA1-79LTK 630KVA-22/0.4KV(ĐH)", "TBA2-79LTK 630KVA-22/0.4KV(AS)",
                        "TBA-33LTT 800KVA-22/0.4KV", "TBA-11THĐ 630KVA-10/0.4KV"]]
    dv_21pdh = df_tonghop.loc["PXHN", "D"]
    dv_baoviet_t5 = df_tonghop.loc["Bảo Việt mới", "D"]
    dv_5ltk = df_tonghop.loc[["TBA1-5LTK 800KVA-22/0.4KV(AS,OC)", "TBA2-5LTK 800KVA-22/0.4KV(ĐH)"], "D"].sum()
    dv_79ltk = df_tonghop.loc[["TBA1-79LTK 630KVA-22/0.4KV(ĐH)", "TBA2-79LTK 630KVA-22/0.4KV(AS)"], "D"].sum()
    dv_33ltt = df_tonghop.loc["TBA-33LTT 800KVA-22/0.4KV", "D"]
    dv_11thd = df_tonghop.loc["TBA-11THĐ 630KVA-10/0.4KV", "D"]

    df_tong_hop_frame = pd.DataFrame({
        'Stt': list(range(1, 22)),
        'Địa chỉ': [
            'Cà phê báo chí tầng 9', 'Nhà ăn tầng 9', 'Sảnh tầng 10', 'Phòng truyền thống tâng 10',
            'Hội trường tầng 10', 'Sảnh tầng 11', 'Màn hình Led hè 5 LTK', 'Báo VN News 79LTK',
            'Số 21PĐH', 'Bảo Việt T5', 'Ngân hàng MB', 'GME', 'Giovanni', 'FSB tầng 12',
            'PG Bank', 'Số 8THĐ', 'Số 5LTK', 'Số 79LTK', 'Số 33LTT', 'Số 11THĐ', 'Tổng'
        ],
        'Sản lượng tuần mới (kWh)': [""] * 21,
        'Sản lượng tuần cũ (kWh)': [""] * 21,
        'Sản lượng tăng (kWh)': [""] * 21
    })

    service_total_list = [df_cafe_t9_frame_total, df_nha_an_t9_frame_total, df_sanh_t10_frame_total, df_phong_truyen_thong_t10_frame_total, df_hoitruong_t10_frame_total,
                        df_t11_frame_total, df_led_frame_total, df_vnnews_frame_total, dv_21pdh, dv_baoviet_t5, df_mb_hoankiem_frame_total, df_gme_frame_total, df_giovani_frame_total,
                        fbs_t12, df_vna8_pgb_frame_total, df_vna8_vp_frame.loc[2, "Thanh toán (KWh)"], dv_5ltk, dv_79ltk, dv_33ltt, dv_11thd]

    df_tong_hop_frame.loc[:19, 'Sản lượng tuần mới (kWh)'] = service_total_list
    df_tong_hop_frame.loc[20, 'Sản lượng tuần mới (kWh)'] = df_tong_hop_frame.loc[16:19, 'Sản lượng tuần mới (kWh)'].sum()
    tong_hop_frame_title = pd.DataFrame([["", "Tuần từ:", f"{df.columns[5]} => {df.columns[6]}", f"{df.columns[4]} => {df.columns[5]}", ""]],
                                        columns=df_tong_hop_frame.columns)
    df_tong_hop_frame = pd.concat([tong_hop_frame_title, df_tong_hop_frame], ignore_index=True)
    tong_hop_sheet_frame = df_tong_hop_frame

    # III. Phần trung bình và tiêu thụ
    # Phần trung bình
    df_tb = df.loc[["TBA1-5LTK 800KVA-22/0.4KV(AS,OC)", "TBA2-5LTK 800KVA-22/0.4KV(ĐH)", "TBA-33LTT 800KVA-22/0.4KV",
                    "TBA-11THĐ 630KVA-10/0.4KV","TBA1-79LTK 630KVA-22/0.4KV(ĐH)", "TBA2-79LTK 630KVA-22/0.4KV(AS)"]]

    df_tb_frame = pd.DataFrame({
        (f"{df.columns[2]} => {df.columns[3]}", 'TB tuần'): [df_tb.iloc[i, 7] for i in range(df_tb.shape[0])]
            + [sum(df_tb.iloc[i, 7] for i in range(df_tb.shape[0]))],

        (f"{df.columns[2]} => {df.columns[3]}", 'TB ngày'): [round(df_tb.iloc[i,7] / 7, 0) for i in range(df_tb.shape[0])]
            + [round(sum(df_tb.iloc[i, 7] for i in range(df_tb.shape[0])) / 7, 0)],

        (f"{df.columns[3]} => {df.columns[4]}", 'TB tuần'): [df_tb.iloc[i, 8] for i in range(df_tb.shape[0])]
            + [sum(df_tb.iloc[i, 8] for i in range(df_tb.shape[0]))],

        (f"{df.columns[3]} => {df.columns[4]}", 'TB ngày'): [round(df_tb.iloc[i, 8] / 7, 0) for i in range(df_tb.shape[0])]
            + [round(sum(df_tb.iloc[i, 8] for i in range(df_tb.shape[0])) / 7, 0)],

        (f"{df.columns[4]} => {df.columns[5]}", 'TB tuần'): [df_tb.iloc[i, 9] for i in range(df_tb.shape[0])]
            + [sum(df_tb.iloc[i, 9] for i in range(df_tb.shape[0]))],

        (f"{df.columns[4]} => {df.columns[5]}", 'TB ngày'): [round(df_tb.iloc[i, 9] / 7, 0) for i in range(df_tb.shape[0])]
            + [round(sum(df_tb.iloc[i, 9] for i in range(df_tb.shape[0])) / 7, 0)],

        (f"{df.columns[5]} => {df.columns[6]}", 'TB tuần'): [df_tb.iloc[i, 10] for i in range(df_tb.shape[0])]
            + [sum(df_tb.iloc[i, 10] for i in range(df_tb.shape[0]))],

        (f"{df.columns[5]} => {df.columns[6]}", 'TB ngày'): [round(df_tb.iloc[i, 10] / 7, 0) for i in range(df_tb.shape[0])]
            + [round(sum(df_tb.iloc[i, 10] for i in range(df_tb.shape[0])) / 7, 0)],
    })

    index = pd.MultiIndex.from_arrays([
        ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII'],
        ['TBA1-5LTK 800KVA-22/0.4KV(AS,OC)', 'TBA2-5LTK 800KVA-22/0.4KV(ĐH)',
        'TBA-33LTT 800KVA-22/0.4KV', 'TBA-11THĐ 630KVA-10/0.4KV',
        'TBA1-79LTK 630KVA-22/0.4KV(ĐH)', 'TBA2-79LTK 630KVA-22/0.4KV(AS)', 'Tổng cộng']
    ], names=['STT', 'Tên công tơ'])

    df_tb_frame.index = index

    # Phần tiêu thụ
    df_tieu_thu_frame = pd.DataFrame({
        (f"{df.columns[2]} => {df.columns[3]}", 'TB tuần'): [""]*5,

        (f"{df.columns[2]} => {df.columns[3]}", 'TB ngày'): [""]*5,

        (f"{df.columns[3]} => {df.columns[4]}", 'TB tuần'): [""]*5,

        (f"{df.columns[3]} => {df.columns[4]}", 'TB ngày'): [""]*5,

        (f"{df.columns[4]} => {df.columns[5]}", 'TB tuần'): [""]*5,

        (f"{df.columns[4]} => {df.columns[5]}", 'TB ngày'): [""]*5,

        (f"{df.columns[5]} => {df.columns[6]}", 'TB tuần'): [""]*5,

        (f"{df.columns[5]} => {df.columns[6]}", 'TB ngày'): [""]*5
    })

    index = pd.MultiIndex.from_arrays([
        ['', 'VIII', 'IX', 'X', 'XI'],
        ['Lượng điện tiêu thụ','Số điện tiêu thụ (VNA8)', 'Số điện tiêu thụ (Dịch vụ)',
        'Số điện tiêu thụ (Cơ quan)', 'Số điện tiêu thụ tổng ']
    ], names=['STT', 'Địa điểm'])

    df_tieu_thu_frame.index = index

    df_tieu_thu_frame.loc[('VIII', 'Số điện tiêu thụ (VNA8)'), (f"{df.columns[3]} => {df.columns[4]}", 'TB tuần')] = df_vna8_total
    df_tieu_thu_frame.loc[('IX', 'Số điện tiêu thụ (Dịch vụ)'), (f"{df.columns[3]} => {df.columns[4]}", 'TB tuần')] = total_service
    df_tieu_thu_frame.loc[('XI', 'Số điện tiêu thụ tổng '), (f"{df.columns[3]} => {df.columns[4]}", 'TB tuần')] = df.loc["Tổng cộng", "D"]
    df_tieu_thu_frame.loc[('X', 'Số điện tiêu thụ (Cơ quan)'), (f"{df.columns[3]} => {df.columns[4]}", 'TB tuần')] = (df.loc["Tổng cộng", "D"] - total_service - df_gme_frame_total)

    df_tieu_thu_frame.loc[('VIII', 'Số điện tiêu thụ (VNA8)'), (f"{df.columns[5]} => {df.columns[6]}", 'TB tuần')] = f"{round((df_vna8_total/df.loc['Tổng cộng', 'D'])*100, 2)}%"
    df_tieu_thu_frame.loc[('IX', 'Số điện tiêu thụ (Dịch vụ)'), (f"{df.columns[5]} => {df.columns[6]}", 'TB tuần')] = f"{round((total_service/df.loc['Tổng cộng', 'D'])*100, 2)}%"
    df_tieu_thu_frame.loc[('XI', 'Số điện tiêu thụ tổng '), (f"{df.columns[5]} => {df.columns[6]}", 'TB tuần')] =  f"{round((df.loc['Tổng cộng', 'D']/df.loc['Tổng cộng', 'D'])*100, 2)}%"
    df_tieu_thu_frame.loc[('X', 'Số điện tiêu thụ (Cơ quan)'), (f"{df.columns[5]} => {df.columns[6]}", 'TB tuần')] = f"{round((df.loc['Tổng cộng', 'D'] - df_vna8_total - total_service)/df.loc['Tổng cộng', 'D']*100, 2)}%"

    df_tieu_thu_frame = pd.concat([df_tb_frame, df_tieu_thu_frame])
    trung_binh_sheet_frame = df_tieu_thu_frame

    return df, dich_vu_sheet_frame, tong_hop_sheet_frame, trung_binh_sheet_frame

def export_excel_file(df, dich_vu_sheet_frame, tong_hop_sheet_frame, trung_binh_sheet_frame):
    file_path = 'bao_cao_so_dien.xlsx'

    # Tạo file Excel với 4 sheet
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Xuất các DataFrame ra các sheet tương ứng
        df.to_excel(writer, sheet_name='Dữ liệu chính', index=True)
        dich_vu_sheet_frame.to_excel(writer, sheet_name='Dịch vụ', index=True)
        tong_hop_sheet_frame.to_excel(writer, sheet_name='Tổng hợp', index=True)
        trung_binh_sheet_frame.to_excel(writer, sheet_name='Trung bình', index=True)

    # Mở file Excel để định dạng
    wb = load_workbook(file_path)

    # Định nghĩa các style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='6fa8dc', end_color='6fa8dc', fill_type='solid')
    alt_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    total_row_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    header_font = Font(bold=True, size=12)
    regular_font = Font(size=11)

    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

    # Áp dụng định dạng cho từng sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Điều chỉnh chiều rộng cột
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Định dạng tiêu đề
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Định dạng dữ liệu
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = thin_border
                cell.font = regular_font

                # Canh lề phù hợp dựa trên loại dữ liệu
                if isinstance(cell.value, (int, float)):
                    cell.alignment = right_alignment
                else:
                    cell.alignment = left_alignment

                # Màu nền cho hàng chẵn/lẻ
                if row_idx % 2 == 0:
                    cell.fill = alt_row_fill

                # Màu nền cho hàng tổng (nếu có)
                if "Tổng" in str(ws.cell(row=row_idx, column=1).value):
                    cell.fill = total_row_fill
                    cell.font = Font(bold=True, size=11)

        # Đóng băng hàng đầu tiên
        ws.freeze_panes = 'A2'

    # Lưu file sau khi định dạng
    wb.save(file_path)

