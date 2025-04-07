import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import xlsxwriter
import io
from bao_cao_tuan import bao_cao_tuan_electric
from bao_cao_thang import bao_cao_thang_electric

def export_excel_formatted_fixed(sheet_df_dict, titles_dict, output_path):
    """
    Export and format dataframes to Excel with proper styling

    Parameters:
    -----------
    sheet_df_dict : dict
        Dictionary with sheet names as keys and lists of dataframes as values
    titles_dict : dict
        Dictionary with sheet names as keys and lists of titles as values
    output_path : str
        Path to save the Excel file
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        header_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        data_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        total_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        thick_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)

        title_font = Font(bold=True, size=14, color='FFFFFF')
        bold_font = Font(bold=True)

        current_row = {}

        for sheet_name, dfs in sheet_df_dict.items():
            titles = titles_dict.get(sheet_name, [])
            current_row[sheet_name] = 0

            if sheet_name not in writer.sheets:
                workbook = writer.book
                workbook.create_sheet(sheet_name)

            for i, dataframe in enumerate(dfs):
                title = titles[i] if i < len(titles) else f"Bảng dữ liệu {i + 1}"

                start_row = current_row[sheet_name] + 1

                # Phát hiện nếu là MultiIndex
                is_multi_index_columns = isinstance(dataframe.columns, pd.MultiIndex)
                is_multi_index_rows = isinstance(dataframe.index, pd.MultiIndex)

                # Tính tổng số cột thực tế
                if is_multi_index_columns:
                    # Duyệt qua các hàng để tìm số cột tối đa
                    actual_max_col = max(len(row) for row in dataframe.values)
                    # +1 cho mỗi cấp độ index
                    max_col = actual_max_col + dataframe.index.nlevels
                else:
                    max_col = len(dataframe.columns) + dataframe.index.nlevels

                # Export DataFrame to Excel
                dataframe.to_excel(writer, sheet_name=sheet_name, startrow=start_row)

                sheet = writer.sheets[sheet_name]

                # Add and format title
                sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)

                title_cell = sheet.cell(row=start_row, column=1)
                title_cell.value = title
                title_cell.alignment = center_alignment
                title_cell.font = title_font
                title_cell.fill = title_fill
                title_cell.border = thin_border

                # Xác định số dòng tiêu đề
                header_rows = dataframe.columns.nlevels if is_multi_index_columns else 1

                # Ẩn dòng trống sau tiêu đề MultiIndex nếu có
                if is_multi_index_columns:
                    blank_row = start_row + 1 + header_rows
                    sheet.row_dimensions[blank_row].hidden = True

                # Format header rows
                for row_offset in range(header_rows):
                    for col in range(1, max_col + 1):
                        header_row = start_row + 1 + row_offset
                        cell = sheet.cell(row=header_row, column=col)
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                        cell.font = bold_font
                        cell.border = thin_border

                # Tìm vị trí các cột đặc biệt
                so_cong_to_col = None
                phan_tram_col = None

                # Duyệt qua các cột header để tìm cột "Số công tơ" và "%"
                for col in range(1, max_col + 1):
                    header_cell = sheet.cell(row=start_row + 1, column=col)
                    header_text = str(header_cell.value).lower() if header_cell.value else ""

                    if "số công tơ" in header_text or "so cong to" in header_text:
                        so_cong_to_col = col
                    elif "%" in header_text:
                        phan_tram_col = col

                # Tính toán phạm vi dòng dữ liệu
                if is_multi_index_columns:
                    data_start_row = start_row + 1 + header_rows + 1  # +1 cho dòng trống
                    data_end_row = start_row + len(dataframe) + header_rows + 1
                else:
                    data_start_row = start_row + 1 + header_rows
                    data_end_row = start_row + len(dataframe) + header_rows

                # Format data rows
                for row in range(data_start_row, data_end_row + 1):
                    # Kiểm tra xem đây có phải dòng tổng
                    is_total_row = False

                    if is_multi_index_rows:
                        idx_col = 1  # Cột đầu tiên (STT)
                        cell_value = str(sheet.cell(row=row, column=idx_col).value).lower()
                        is_total_row = "tổng" in cell_value or "vii" in cell_value or "total" in cell_value
                    else:
                        cell_value = str(sheet.cell(row=row, column=2).value).lower() if sheet.cell(row=row,
                                                                                                    column=2).value else ""
                        is_total_row = "tổng" in cell_value or "total" in cell_value

                    for col in range(1, max_col + 1):
                        cell = sheet.cell(row=row, column=col)

                        if is_total_row:
                            cell.fill = total_fill
                            cell.font = bold_font
                        else:
                            cell.fill = data_fill

                        cell.border = thin_border

                        # Áp dụng định dạng đặc biệt cho các cột
                        if isinstance(cell.value, (int, float)):
                            # Định dạng cột "Số công tơ" - số nguyên không dấu phẩy
                            if so_cong_to_col and col == so_cong_to_col:
                                cell.number_format = '0'
                                cell.alignment = center_alignment
                            # Định dạng cột "%" - 2 chữ số sau dấu phẩy
                            elif phan_tram_col and col == phan_tram_col:
                                cell.number_format = '0.00'
                                cell.alignment = right_alignment
                            # Định dạng các cột số khác
                            elif col > dataframe.index.nlevels:
                                cell.number_format = '#,##0'
                                cell.alignment = right_alignment
                        else:
                            cell.alignment = center_alignment

                # Thêm viền dày xung quanh bảng
                # Viền trên
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=start_row, column=col)
                    cell.border = Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=Side(style='medium'),
                        bottom=cell.border.bottom
                    )

                # Viền dưới
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=data_end_row, column=col)
                    cell.border = Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=Side(style='medium')
                    )

                # Viền trái phải
                for row in range(start_row, data_end_row + 1):
                    # Viền trái
                    cell = sheet.cell(row=row, column=1)
                    cell.border = Border(
                        left=Side(style='medium'),
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=cell.border.bottom
                    )

                    # Viền phải - QUAN TRỌNG: Đảm bảo áp dụng cho cột cuối cùng thực tế
                    cell = sheet.cell(row=row, column=max_col)
                    cell.border = Border(
                        left=cell.border.left,
                        right=Side(style='medium'),
                        top=cell.border.top,
                        bottom=cell.border.bottom
                    )

                # Điều chỉnh vị trí dòng bắt đầu tiếp theo
                current_row[sheet_name] = data_end_row + 2
